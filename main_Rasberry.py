# coding: utf-8
import cv2
import numpy as np
import pandas as pd
import torch
from datetime import datetime
from PIL import Image
from facenet_pytorch import MTCNN, InceptionResnetV1
import joblib
from sklearn.preprocessing import LabelEncoder
import os
import time
import atexit
from RPLCD.i2c import CharLCD
from email.message import EmailMessage
import smtplib
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Initialiser LCD
lcd = CharLCD('PCF8574', 0x27)

def initialize_components():
    device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
    mtcnn = MTCNN(keep_all=True, device=device)
    facenet = InceptionResnetV1(pretrained='vggface2').eval().to(device)

    try:
        classifier = joblib.load('face_classifiersa.pkl')
        label_encoder = LabelEncoder()
        label_encoder.classes_ = np.load('label_encodersa.npy', allow_pickle=True)
    except Exception as e:
        print(f"Erreur lors du chargement des modèles: {e}")
        afficher_sur_lcd("Erreur Modeles", "Redemarrez")
        exit()

    return device, mtcnn, facenet, classifier, label_encoder

def get_embedding(face_pixels, model, device):
    face_tensor = torch.tensor(face_pixels.transpose(2, 0, 1)).unsqueeze(0).float().to(device)
    with torch.no_grad():
        embedding = model(face_tensor)
    return embedding.cpu().numpy().reshape(-1)

def afficher_sur_lcd(ligne1, ligne2=""):
    lcd.clear()
    lcd.write_string(ligne1[:16])
    if ligne2:
        lcd.cursor_pos = (1, 0)
        lcd.write_string(ligne2[:16])

def envoyer_email_avec_piece_jointe(fichier_piece_jointe, destinataire, sujet, message):
    EMAIL_EXPEDITEUR = "soadsalma1234321@gmail.com"
    MOT_DE_PASSE = "vhnmsynlwullqidh"

    if not os.path.exists(fichier_piece_jointe):
        print(f"❌ Le fichier {fichier_piece_jointe} est introuvable.")
        return

    msg = EmailMessage()
    msg['Subject'] = sujet
    msg['From'] = EMAIL_EXPEDITEUR
    msg['To'] = destinataire
    msg.set_content(message)

    with open(fichier_piece_jointe, 'rb') as f:
        contenu = f.read()
        msg.add_attachment(
            contenu,
            maintype='application',
            subtype='octet-stream',
            filename=os.path.basename(fichier_piece_jointe)
        )

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(EMAIL_EXPEDITEUR, MOT_DE_PASSE)
            smtp.send_message(msg)
            print("✅ Email envoyé avec succès à", destinataire)
        afficher_sur_lcd("Mail envoyé", "au professeur")
    except Exception as e:
        print(f"❌ Erreur lors de l'envoi de l'email : {e}")
        afficher_sur_lcd("Erreur envoi", "mail")

def main():
    device, mtcnn, facenet, classifier, label_encoder = initialize_components()
    DETECTION_DURATION = 180  # Durée totale de détection (3 minutes)
    MIN_DETECTION_TIME = 7    # 7 secondes de détection continue
    TAILLE_LIMITE = 200       # Taille minimale du visage détecté
    attendance_file = "liste_presence.xlsx"
    unknown_path = "inconnus"
    cap = cv2.VideoCapture(0)

    # Nettoyage à la fermeture
    def cleanup():
        """Supprime les fichiers temporaires"""
        for f in os.listdir(unknown_path):
            if f.startswith('temp_'):
                os.remove(os.path.join(unknown_path, f))
    atexit.register(cleanup)

    if not cap.isOpened():
        print("Erreur: Impossible d'ouvrir la caméra")
        return

    # Initialisation Excel
    wb = Workbook()
    ws = wb.active
    ws.append(["Nom", "Prénom", "Statut", "Heure_detection", "Photo"])
    
    # Configuration des colonnes
    ws.column_dimensions['A'].width = 15  # Nom
    ws.column_dimensions['B'].width = 15  # Prénom
    ws.column_dimensions['C'].width = 15  # Statut
    ws.column_dimensions['D'].width = 20  # Heure
    ws.column_dimensions['E'].width = 30  # Photo
    
    IMAGE_ROW_HEIGHT = 120  # Hauteur pour les images

    detection_history = {
        label: {'detected': False, 'time': None, 'announced': False}
        for label in label_encoder.classes_
    }

    unknown_faces_cache = []
    unknown_counter = 1
    os.makedirs(unknown_path, exist_ok=True)

    # Variables pour le suivi temporel
    current_face_start_time = None
    current_face_label = None
    current_face_embedding = None
    current_face_box = None

    afficher_sur_lcd("Systeme actif", "En attente...")
    print("En attente de détection d'une personne proche...")

    start_time = datetime.now()
    try:
        while True:
            elapsed_time = (datetime.now() - start_time).total_seconds()
            if elapsed_time > DETECTION_DURATION:
                afficher_sur_lcd("Temps terminé", "Fin détection")
                time.sleep(3)
                break

            ret, frame = cap.read()
            if not ret:
                continue

            img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
            boxes, _ = mtcnn.detect(img)

            if boxes is not None:
                max_hauteur = 0
                box_proche = None
                for box in boxes:
                    hauteur = box[3] - box[1]
                    if hauteur > max_hauteur and hauteur > TAILLE_LIMITE:
                        max_hauteur = hauteur
                        box_proche = box

                if box_proche is not None:
                    # Si nouveau visage ou différent du précédent
                    face = img.crop(box_proche)
                    face = face.resize((160, 160))
                    face_array = np.array(face).astype('float32')
                    face_array = (face_array - 127.5) / 128.0
                    embedding = get_embedding(face_array, facenet, device)
                    proba = classifier.predict_proba(embedding.reshape(1, -1))
                    max_proba = np.max(proba)
                    
                    if max_proba > 0.7:
                        new_label = label_encoder.inverse_transform([np.argmax(proba)])[0]
                    else:
                        new_label = "inconnu"
                    
                    # Si visage différent du précédent
                    if current_face_label != new_label or current_face_start_time is None:
                        current_face_start_time = datetime.now()
                        current_face_label = new_label
                        current_face_embedding = embedding
                        current_face_box = box_proche
                        afficher_sur_lcd("Detection en", "cours...")
                    
                    # Vérifier si le visage est présent depuis 7 secondes
                    detection_duration = (datetime.now() - current_face_start_time).total_seconds()
                    if detection_duration >= MIN_DETECTION_TIME:
                        if current_face_label != "inconnu":
                            # Visage reconnu
                            nom, prenom = current_face_label.split('_')
                            if not detection_history[current_face_label]['detected']:
                                detection_history[current_face_label]['detected'] = True
                                detection_history[current_face_label]['time'] = datetime.now()
                                afficher_sur_lcd("Présence:", f"{nom} {prenom}")
                                ws.append([nom, prenom, "Présent", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
                                time.sleep(3)
                        else:
                            # Visage inconnu
                            is_new_unknown = True
                            for known_emb in unknown_faces_cache:
                                distance = np.linalg.norm(embedding - known_emb)
                                if distance < 0.8:
                                    is_new_unknown = False
                                    break

                            if is_new_unknown:
                                afficher_sur_lcd("Nouveau visage", "Capture en cours...")
                                time.sleep(1)
                                
                                # Sauvegarde image
                                unknown_name = f"Inconnu_{unknown_counter}"
                                filename = f"{unknown_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                                save_path = os.path.join(unknown_path, filename)
                                face_bgr = cv2.cvtColor(np.array(face), cv2.COLOR_RGB2BGR)
                                cv2.imwrite(save_path, face_bgr)
                                
                                # Insertion dans Excel avec image
                                ws.append([unknown_name, "", "Inconnu", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
                                current_row = ws.max_row
                                ws.row_dimensions[current_row].height = IMAGE_ROW_HEIGHT
                                
                                try:
                                    img_excel = ExcelImage(save_path)
                                    img_excel.width = 150
                                    img_excel.height = 120
                                    ws.add_image(img_excel, f'E{current_row}')
                                except Exception as e:
                                    print(f"Erreur insertion image: {e}")
                                    afficher_sur_lcd("Erreur image", "Reessayer")
                                
                                afficher_sur_lcd("Capture", "Enregistrée!")
                                unknown_faces_cache.append(embedding)
                                unknown_counter += 1
                                time.sleep(2)
                        
                        # Réinitialiser pour la prochaine détection
                        current_face_start_time = None
                        current_face_label = None
                else:
                    # Pas de visage détecté
                    current_face_start_time = None
                    current_face_label = None
            else:
                # Pas de visage détecté
                current_face_start_time = None
                current_face_label = None

            if os.environ.get('DISPLAY'):
                cv2.imshow('Reconnaissance Faciale', frame)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break

    finally:
        # Ajouter les absents
        current_time = datetime.now()
        for label in detection_history:
            nom, prenom = label.split('_')
            if not detection_history[label]['detected']:
                ws.append([nom, prenom, "Absent", "N/A", ""])

        # Sauvegarder Excel
        wb.save(attendance_file)
        cap.release()
        cv2.destroyAllWindows()
        lcd.clear()
        afficher_sur_lcd("Détection finie", "Excel enregistré")

        # Envoyer email
        envoyer_email_avec_piece_jointe(
            fichier_piece_jointe=attendance_file,
            destinataire="salmabourkiba2021@gmail.com",
            sujet="Liste de présence IA - Reconnaissance faciale",
            message="Bonjour,\n\nVeuillez trouver ci-joint la liste de présence générée automatiquement par le système.\n\nCordialement,"
        )

if __name__ == "__main__":
    required_files = ['face_classifiersa.pkl', 'label_encodersa.npy']
    missing_files = [f for f in required_files if not os.path.exists(f)]

    if missing_files:
        print("Fichiers manquants :")
        for f in missing_files:
            print(f"- {f}")
        afficher_sur_lcd("Erreur:", "Fichiers manquants")
    else:
        main()